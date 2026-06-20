from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib import messages
from django.http import HttpResponse

from django.contrib.auth.models import User

from openpyxl import Workbook


from .models import CourseSchedule, TimeSlot, Teacher, Department, Section, Level, Subject, Room
from .forms import CourseScheduleForm, TeacherForm, DepartmentForm, SectionForm, LevelForm, SubjectForm, RoomForm, ChiefUserForm


def is_admin(user):
    return user.is_staff or user.is_superuser


@login_required
def dashboard(request):
    if request.user.is_staff or request.user.is_superuser:
        schedules_count = CourseSchedule.objects.count()
    else:
        try:
            section = request.user.managed_section
            schedules_count = CourseSchedule.objects.filter(
                level__section=section
            ).count()
        except:
            schedules_count = 0

    return render(
        request,
        "timetable/dashboard.html",
        {"schedules_count": schedules_count}
    )


@login_required
def timetable_list(request):
    if request.user.is_staff or request.user.is_superuser:
        schedules = CourseSchedule.objects.select_related(
            "subject",
            "teacher",
            "room",
            "level",
            "time_slot"
        )
    else:
        section = request.user.managed_section
        schedules = CourseSchedule.objects.filter(
            level__section=section
        ).select_related(
            "subject",
            "teacher",
            "room",
            "level",
            "time_slot"
        )

    return render(
        request,
        "timetable/timetable_list.html",
        {"schedules": schedules}
    )


@login_required
def timetable_grid(request):
    days = [
        "LUNDI",
        "MARDI",
        "MERCREDI",
        "JEUDI",
        "VENDREDI",
        "SAMEDI",
    ]

    time_slots = TimeSlot.objects.all().order_by("start_time")

    if request.user.is_staff or request.user.is_superuser:
        schedules = CourseSchedule.objects.select_related(
            "subject",
            "teacher",
            "room",
            "level",
            "time_slot"
        )
    else:
        section = request.user.managed_section
        schedules = CourseSchedule.objects.filter(
            level__section=section
        ).select_related(
            "subject",
            "teacher",
            "room",
            "level",
            "time_slot"
        )

    timetable = {}

    for day in days:
        timetable[day] = {}

        for slot in time_slots:
            timetable[day][slot] = None

    for course in schedules:
        timetable[course.day][course.time_slot] = course

    return render(
        request,
        "timetable/timetable_grid.html",
        {
            "timetable": timetable,
            "time_slots": time_slots
        }
    )


@login_required
def course_create(request):
    if request.method == "POST":
        form = CourseScheduleForm(
            request.POST,
            user=request.user
        )

        if form.is_valid():
            course = form.save(commit=False)
            course.created_by = request.user
            course.save()

            messages.success(
                request,
                "Cours ajouté avec succès."
            )

            return redirect("timetable:timetable_grid")

    else:
        form = CourseScheduleForm(user=request.user)

    return render(
        request,
        "timetable/course_form.html",
        {"form": form}
    )


@login_required
def course_update(request, pk):
    course = get_object_or_404(CourseSchedule, pk=pk)

    if not request.user.is_staff and not request.user.is_superuser:
        if course.level.section != request.user.managed_section:
            messages.error(request, "Accès refusé.")
            return redirect("timetable:timetable_grid")

    if request.method == "POST":
        form = CourseScheduleForm(
            request.POST,
            instance=course,
            user=request.user
        )

        if form.is_valid():
            form.save()

            messages.success(
                request,
                "Cours modifié avec succès."
            )

            return redirect("timetable:timetable_grid")

    else:
        form = CourseScheduleForm(
            instance=course,
            user=request.user
        )

    return render(
        request,
        "timetable/course_form.html",
        {"form": form}
    )


@login_required
def course_delete(request, pk):
    course = get_object_or_404(CourseSchedule, pk=pk)

    if not request.user.is_staff and not request.user.is_superuser:
        if course.level.section != request.user.managed_section:
            messages.error(request, "Accès refusé.")
            return redirect("timetable:timetable_grid")

    if request.method == "POST":
        course.delete()

        messages.success(
            request,
            "Cours supprimé."
        )

        return redirect("timetable:timetable_grid")

    return render(
        request,
        "timetable/confirm_delete.html",
        {"object": course}
    )


@login_required
def change_password(request):
    if request.method == "POST":
        form = PasswordChangeForm(
            request.user,
            request.POST
        )

        if form.is_valid():
            user = form.save()

            update_session_auth_hash(
                request,
                user
            )

            messages.success(
                request,
                "Mot de passe modifié."
            )

            return redirect("timetable:dashboard")

    else:
        form = PasswordChangeForm(request.user)

    return render(
        request,
        "timetable/change_password.html",
        {"form": form}
    )


@login_required
@user_passes_test(is_admin)
def reset_chief_password(request, user_id):
    from django.contrib.auth.models import User

    user = get_object_or_404(User, pk=user_id)

    user.set_password("1234")
    user.save()

    messages.success(
        request,
        f"Mot de passe de {user.username} réinitialisé à 1234."
    )

    return redirect("admin:auth_user_changelist")


@login_required
def export_timetable_excel(request):
    if request.user.is_staff or request.user.is_superuser:
        schedules = CourseSchedule.objects.all().select_related(
            "subject",
            "teacher",
            "room",
            "level",
            "time_slot"
        )
    else:
        section = request.user.managed_section
        schedules = CourseSchedule.objects.filter(
            level__section=section
        ).select_related(
            "subject",
            "teacher",
            "room",
            "level",
            "time_slot"
        )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Emploi du temps"

    headers = [
        "Jour",
        "Créneau",
        "Classe",
        "Matière",
        "Enseignant",
        "Salle"
    ]

    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num).value = header

    row = 2

    for schedule in schedules:
        sheet.cell(row=row, column=1).value = schedule.day
        sheet.cell(row=row, column=2).value = str(schedule.time_slot)
        sheet.cell(row=row, column=3).value = str(schedule.level)
        sheet.cell(row=row, column=4).value = str(schedule.subject)
        sheet.cell(row=row, column=5).value = str(schedule.teacher)
        sheet.cell(row=row, column=6).value = str(schedule.room)

        row += 1

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = (
        'attachment; filename="emploi_du_temps.xlsx"'
    )

    workbook.save(response)

    return response



@login_required
def teacher_list(request):

    teachers = Teacher.objects.all().order_by(
        "last_name",
        "first_name"
    )

    return render(
        request,
        "timetable/teacher_list.html",
        {"teachers": teachers}
    )


@login_required
@user_passes_test(is_admin)
def teacher_create(request):

    if request.method == "POST":

        form = TeacherForm(request.POST)

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Enseignant ajouté avec succès."
            )

            return redirect(
                "timetable:teacher_list"
            )

    else:

        form = TeacherForm()

    return render(
        request,
        "timetable/teacher_form.html",
        {"form": form}
    )


@login_required
@user_passes_test(is_admin)
def teacher_update(request, pk):

    teacher = get_object_or_404(
        Teacher,
        pk=pk
    )

    if request.method == "POST":

        form = TeacherForm(
            request.POST,
            instance=teacher
        )

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Enseignant modifié."
            )

            return redirect(
                "timetable:teacher_list"
            )

    else:

        form = TeacherForm(
            instance=teacher
        )

    return render(
        request,
        "timetable/teacher_form.html",
        {"form": form}
    )


@login_required
@user_passes_test(is_admin)
def teacher_delete(request, pk):

    teacher = get_object_or_404(
        Teacher,
        pk=pk
    )

    if request.method == "POST":

        teacher.delete()

        messages.success(
            request,
            "Enseignant supprimé."
        )

        return redirect(
            "timetable:teacher_list"
        )

    return render(
        request,
        "timetable/confirm_delete.html",
        {"object": teacher}
    )


@login_required
def department_list(request):

    departments = Department.objects.all().order_by("name")

    return render(
        request,
        "timetable/department_list.html",
        {"departments": departments}
    )


@login_required
@user_passes_test(is_admin)
def department_create(request):

    if request.method == "POST":

        form = DepartmentForm(request.POST)

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Département ajouté avec succès."
            )

            return redirect(
                "timetable:department_list"
            )

    else:

        form = DepartmentForm()

    return render(
        request,
        "timetable/department_form.html",
        {"form": form}
    )
@login_required
@user_passes_test(is_admin)
def teacher_delete(request, pk):
    teacher = get_object_or_404(Teacher, pk=pk)

    if request.method == "POST":
        teacher.delete()
        messages.success(request, "Enseignant supprimé.")
        return redirect("timetable:teacher_list")

    return render(
        request,
        "timetable/confirm_delete.html",
        {"object": teacher}
    )



@login_required
@user_passes_test(is_admin)
def department_update(request, pk):

    department = get_object_or_404(
        Department,
        pk=pk
    )

    if request.method == "POST":

        form = DepartmentForm(
            request.POST,
            instance=department
        )

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Département modifié."
            )

            return redirect(
                "timetable:department_list"
            )

    else:

        form = DepartmentForm(
            instance=department
        )

    return render(
        request,
        "timetable/department_form.html",
        {"form": form}
    )


@login_required
@user_passes_test(is_admin)
def department_delete(request, pk):

    department = get_object_or_404(
        Department,
        pk=pk
    )

    if request.method == "POST":

        department.delete()

        messages.success(
            request,
            "Département supprimé."
        )

        return redirect(
            "timetable:department_list"
        )

    return render(
        request,
        "timetable/confirm_delete.html",
        {"object": department}
    )




@login_required
def section_list(request):

    sections = Section.objects.select_related(
        "department",
        "chief"
    )

    return render(
        request,
        "timetable/section_list.html",
        {"sections": sections}
    )


@login_required
@user_passes_test(is_admin)
def section_create(request):

    if request.method == "POST":

        form = SectionForm(request.POST)

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Section ajoutée avec succès."
            )

            return redirect(
                "timetable:section_list"
            )

    else:

        form = SectionForm()

    return render(
        request,
        "timetable/section_form.html",
        {"form": form}
    )


@login_required
@user_passes_test(is_admin)
def section_update(request, pk):

    section = get_object_or_404(
        Section,
        pk=pk
    )

    if request.method == "POST":

        form = SectionForm(
            request.POST,
            instance=section
        )

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Section modifiée."
            )

            return redirect(
                "timetable:section_list"
            )

    else:

        form = SectionForm(
            instance=section
        )

    return render(
        request,
        "timetable/section_form.html",
        {"form": form}
    )


@login_required
@user_passes_test(is_admin)
def section_delete(request, pk):

    section = get_object_or_404(
        Section,
        pk=pk
    )

    if request.method == "POST":

        section.delete()

        messages.success(
            request,
            "Section supprimée."
        )

        return redirect(
            "timetable:section_list"
        )

    return render(
        request,
        "timetable/confirm_delete.html",
        {"object": section}
    )




@login_required
def level_list(request):

    levels = Level.objects.select_related(
        "section",
        "section__department"
    ).order_by(
        "section__department__name",
        "section__name",
        "name"
    )

    return render(
        request,
        "timetable/level_list.html",
        {"levels": levels}
    )


@login_required
@user_passes_test(is_admin)
def level_create(request):

    if request.method == "POST":

        form = LevelForm(request.POST)

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Niveau ajouté avec succès."
            )

            return redirect("timetable:level_list")

    else:

        form = LevelForm()

    return render(
        request,
        "timetable/level_form.html",
        {"form": form}
    )


@login_required
@user_passes_test(is_admin)
def level_update(request, pk):

    level = get_object_or_404(Level, pk=pk)

    if request.method == "POST":

        form = LevelForm(
            request.POST,
            instance=level
        )

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Niveau modifié."
            )

            return redirect("timetable:level_list")

    else:

        form = LevelForm(instance=level)

    return render(
        request,
        "timetable/level_form.html",
        {"form": form}
    )


@login_required
@user_passes_test(is_admin)
def level_delete(request, pk):

    level = get_object_or_404(Level, pk=pk)

    if request.method == "POST":

        level.delete()

        messages.success(
            request,
            "Niveau supprimé."
        )

        return redirect("timetable:level_list")

    return render(
        request,
        "timetable/confirm_delete.html",
        {"object": level}
    )




@login_required
def room_list(request):

    rooms = Room.objects.all().order_by("name")

    return render(
        request,
        "timetable/room_list.html",
        {"rooms": rooms}
    )


@login_required
@user_passes_test(is_admin)
def room_create(request):

    if request.method == "POST":

        form = RoomForm(request.POST)

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Salle ajoutée avec succès."
            )

            return redirect("timetable:room_list")

    else:

        form = RoomForm()

    return render(
        request,
        "timetable/room_form.html",
        {"form": form}
    )


@login_required
@user_passes_test(is_admin)
def room_update(request, pk):

    room = get_object_or_404(Room, pk=pk)

    if request.method == "POST":

        form = RoomForm(
            request.POST,
            instance=room
        )

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Salle modifiée."
            )

            return redirect("timetable:room_list")

    else:

        form = RoomForm(instance=room)

    return render(
        request,
        "timetable/room_form.html",
        {"form": form}
    )


@login_required
@user_passes_test(is_admin)
def room_delete(request, pk):

    room = get_object_or_404(Room, pk=pk)

    if request.method == "POST":

        room.delete()

        messages.success(
            request,
            "Salle supprimée."
        )

        return redirect("timetable:room_list")

    return render(
        request,
        "timetable/confirm_delete.html",
        {"object": room}
    )






@login_required
def subject_list(request):

    subjects = Subject.objects.select_related(
        "level",
        "level__section"
    ).order_by("name")

    return render(
        request,
        "timetable/subject_list.html",
        {"subjects": subjects}
    )


@login_required
@user_passes_test(is_admin)
def subject_create(request):

    if request.method == "POST":

        form = SubjectForm(request.POST)

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Matière ajoutée."
            )

            return redirect(
                "timetable:subject_list"
            )

    else:

        form = SubjectForm()

    return render(
        request,
        "timetable/subject_form.html",
        {"form": form}
    )


@login_required
@user_passes_test(is_admin)
def subject_update(request, pk):

    subject = get_object_or_404(
        Subject,
        pk=pk
    )

    if request.method == "POST":

        form = SubjectForm(
            request.POST,
            instance=subject
        )

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Matière modifiée."
            )

            return redirect(
                "timetable:subject_list"
            )

    else:

        form = SubjectForm(
            instance=subject
        )

    return render(
        request,
        "timetable/subject_form.html",
        {"form": form}
    )


@login_required
@user_passes_test(is_admin)
def subject_delete(request, pk):

    subject = get_object_or_404(
        Subject,
        pk=pk
    )

    if request.method == "POST":

        subject.delete()

        messages.success(
            request,
            "Matière supprimée."
        )

        return redirect(
            "timetable:subject_list"
        )

    return render(
        request,
        "timetable/confirm_delete.html",
        {"object": subject}
    )



@login_required
@user_passes_test(is_admin)
def chief_list(request):

    chiefs = User.objects.filter(
        managed_section__isnull=False
    ).select_related(
        "managed_section"
    )

    return render(
        request,
        "timetable/chief_list.html",
        {"chiefs": chiefs}
    )


@login_required
@user_passes_test(is_admin)
def chief_create(request):

    if request.method == "POST":

        form = ChiefUserForm(request.POST)

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Chef créé avec succès."
            )

            return redirect(
                "timetable:chief_list"
            )

    else:

        form = ChiefUserForm(
            initial={"password": "1234"}
        )

    return render(
        request,
        "timetable/chief_form.html",
        {"form": form}
    )