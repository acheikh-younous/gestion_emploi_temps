import os

from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from timetable.models import CourseSchedule, Section
from timetable.views.schedule_views import get_selected_week


def format_course(course):
    if not course:
        return "Libre"

    return (
        f"{course.subject.name}\n"
        f"{course.teacher}\n"
        f"{course.subject.level.name}\n"
        f"Salle : {course.room.name}"
    )


@login_required
def export_timetable_excel(request):
    days = ["LUNDI", "MARDI", "MERCREDI", "JEUDI", "VENDREDI", "SAMEDI"]

    columns = [
        "Jour",
        "07h30 - 09h30",
        "09h30 - 09h45",
        "09h45 - 12h45",
        "12h45 - 14h00",
        "14h00 - 17h00",
    ]

    selected_week = get_selected_week(request)

    if request.user.is_staff or request.user.is_superuser:
        sections = Section.objects.select_related("department", "chief").order_by(
            "department__name",
            "name",
        )
    else:
        sections = Section.objects.filter(
            id=request.user.managed_section.id
        ).select_related("department", "chief")

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    for section in sections:
        sheet = workbook.create_sheet(title=section.name[:31])

        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
        sheet.page_margins.left = 0.3
        sheet.page_margins.right = 0.3
        sheet.page_margins.top = 0.5
        sheet.page_margins.bottom = 0.5

        logo_path = os.path.join(
            settings.BASE_DIR,
            "timetable",
            "static",
            "timetable",
            "images",
            "logo.png",
        )

        if os.path.exists(logo_path):
            logo = Image(logo_path)
            logo.width = 80
            logo.height = 80
            sheet.add_image(logo, "A1")

        sheet.merge_cells("B1:F1")
        sheet["B1"] = "EMPLOI DU TEMPS HEBDOMADAIRE"
        sheet["B1"].font = Font(bold=True, size=16)
        sheet["B1"].alignment = Alignment(horizontal="center")

        sheet.merge_cells("B2:F2")
        sheet["B2"] = f"Département : {section.department.name}"
        sheet["B2"].font = Font(bold=True)
        sheet["B2"].alignment = Alignment(horizontal="center")

        sheet.merge_cells("B3:F3")
        sheet["B3"] = f"Section : {section.name}"
        sheet["B3"].font = Font(bold=True)
        sheet["B3"].alignment = Alignment(horizontal="center")

        sheet.merge_cells("B4:F4")
        if selected_week:
            sheet["B4"] = (
                f"Semaine : {selected_week.title} | "
                f"{selected_week.start_date} au {selected_week.end_date}"
            )
        else:
            sheet["B4"] = "Semaine : non définie"

        sheet["B4"].alignment = Alignment(horizontal="center")

        start_row = 6

        for col_num, header in enumerate(columns, 1):
            cell = sheet.cell(row=start_row, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="9C6B30")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        schedules = CourseSchedule.objects.filter(
            subject__level__section=section
        ).select_related(
            "week",
            "subject",
            "subject__level",
            "teacher",
            "room",
            "time_slot",
        )

        if selected_week:
            schedules = schedules.filter(week=selected_week)

        row = start_row + 1

        for day in days:
            sheet.cell(row=row, column=1).value = day
            day_courses = schedules.filter(day=day)

            slot_1 = day_courses.filter(time_slot__start_time="07:30").first()
            slot_2 = day_courses.filter(time_slot__start_time="09:45").first()
            slot_3 = day_courses.filter(time_slot__start_time="14:00").first()

            sheet.cell(row=row, column=2).value = format_course(slot_1)
            sheet.cell(row=row, column=3).value = "PETITE PAUSE"
            sheet.cell(row=row, column=4).value = format_course(slot_2)
            sheet.cell(row=row, column=5).value = "GRANDE PAUSE"
            sheet.cell(row=row, column=6).value = format_course(slot_3)

            row += 1

        signature_row = row + 3

        sheet.merge_cells(
            start_row=signature_row,
            start_column=5,
            end_row=signature_row,
            end_column=6,
        )
        sheet.cell(row=signature_row, column=5).value = "Le Chef de section"
        sheet.cell(row=signature_row, column=5).alignment = Alignment(
            horizontal="center"
        )

        sheet.merge_cells(
            start_row=signature_row + 3,
            start_column=5,
            end_row=signature_row + 3,
            end_column=6,
        )

        chief_name = (
            f"{section.chief.first_name} {section.chief.last_name}"
            if section.chief
            else "Non défini"
        )

        sheet.cell(row=signature_row + 3, column=5).value = chief_name
        sheet.cell(row=signature_row + 3, column=5).font = Font(bold=True)
        sheet.cell(row=signature_row + 3, column=5).alignment = Alignment(
            horizontal="center"
        )

        thin = Side(border_style="thin", color="000000")

        for rows in sheet.iter_rows(
            min_row=start_row,
            max_row=row - 1,
            min_col=1,
            max_col=6,
        ):
            for cell in rows:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )

        for col in range(1, 7):
            sheet.column_dimensions[get_column_letter(col)].width = 25

        for r in range(start_row + 1, row):
            sheet.row_dimensions[r].height = 75

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    )
    response["Content-Disposition"] = (
        'attachment; filename="emplois_du_temps_par_section.xlsx"'
    )

    workbook.save(response)
    return response