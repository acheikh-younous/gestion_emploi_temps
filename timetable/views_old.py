import os

from django.conf import settings
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib import messages
from django.http import HttpResponse
from django.contrib.auth.models import User

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image

from .models import (
    CourseSchedule,
    TimeSlot,
    Teacher,
    Department,
    Section,
    Level,
    Subject,
    Room,
    AcademicWeek,
)

from .forms import (
    CourseScheduleForm,
    TeacherForm,
    DepartmentForm,
    SectionForm,
    LevelForm,
    SubjectForm,
    RoomForm,
    ChiefUserForm,
    AcademicWeekForm,
)


def is_admin(user):
    return user.is_staff or user.is_superuser


def get_user_section(user):
    if user.is_staff or user.is_superuser:
        return None

    return user.managed_section


def user_can_manage_subject(user, subject):
    if user.is_staff or user.is_superuser:
        return True

    return subject.level.section == user.managed_section


def user_can_manage_course(user, course):
    if user.is_staff or user.is_superuser:
        return True

    return course.subject.level.section == user.managed_section


def get_selected_week(request):
    week_id = request.GET.get("week")

    if week_id:
        week = AcademicWeek.objects.filter(id=week_id).first()
        if week:
            return week

    week = AcademicWeek.objects.filter(is_active=True).order_by("-start_date").first()

    if not week:
        week = AcademicWeek.objects.order_by("-start_date").first()

    return week


@login_required
def dashboard(request):
    if request.user.is_staff or request.user.is_superuser:
        context = {
            "departments_count": Department.objects.count(),
            "sections_count": Section.objects.count(),
            "levels_count": Level.objects.count(),
            "teachers_count": Teacher.objects.count(),
            "subjects_count": Subject.objects.count(),
            "rooms_count": Room.objects.count(),
            "weeks_count": AcademicWeek.objects.count(),
            "schedules_count": CourseSchedule.objects.count(),
        }
    else:
        section = request.user.managed_section

        context = {
            "teachers_count": Teacher.objects.count(),
            "subjects_count": Subject.objects.filter(
                level__section=section
            ).count(),
            "levels_count": Level.objects.filter(
                section=section
            ).count(),
            "schedules_count": CourseSchedule.objects.filter(
                subject__level__section=section
            ).count(),
        }

    return render(request, "timetable/dashboard.html", context)


@login_required
def timetable_list(request):
    selected_week = get_selected_week(request)

    schedules = CourseSchedule.objects.select_related(
        "week",
        "subject",
        "subject__level",
        "subject__level__section",
        "teacher",
        "room",
        "time_slot",
    )

    if selected_week:
        schedules = schedules.filter(week=selected_week)

    if not request.user.is_staff and not request.user.is_superuser:
        section = request.user.managed_section
        schedules = schedules.filter(subject__level__section=section)

    schedules = schedules.order_by(
        "week__start_date",
        "subject__level__section__name",
        "subject__level__name",
        "day",
        "time_slot__start_time",
    )

    weeks = AcademicWeek.objects.all().order_by("-start_date")

    return render(
        request,
        "timetable/timetable_list.html",
        {
            "schedules": schedules,
            "weeks": weeks,
            "selected_week": selected_week,
        }
    )


@login_required
def timetable_grid(request):
    days = [
        "LUNDI",
        "MARDI",
        "MERCREDI",
        "JEUDI",
        "VENDREDI",
        "SAMEDI",
    ]

    selected_week = get_selected_week(request)
    weeks = AcademicWeek.objects.all().order_by("-start_date")
    work_slots = TimeSlot.objects.all().order_by("start_time")

    schedules = CourseSchedule.objects.select_related(
        "week",
        "subject",
        "subject__level",
        "subject__level__section",
        "teacher",
        "room",
        "time_slot",
    )

    if selected_week:
        schedules = schedules.filter(week=selected_week)

    if request.user.is_staff or request.user.is_superuser:
        section_id = request.GET.get("section")

        if section_id:
            schedules = schedules.filter(
                subject__level__section_id=section_id
            )

        sections = Section.objects.select_related("department").order_by(
            "department__name",
            "name"
        )

    else:
        section = request.user.managed_section
        schedules = schedules.filter(
            subject__level__section=section
        )
        sections = None
        section_id = section.id

    timetable = {}

    for day in days:
        row = []

        for slot in work_slots:
            course = schedules.filter(
                day=day,
                time_slot=slot
            ).first()

            row.append({
                "type": "course",
                "slot": slot,
                "course": course
            })

            if slot.start_time.strftime("%H:%M") == "07:30":
                row.append({
                    "type": "pause",
                    "label": "Petite pause"
                })

            if slot.start_time.strftime("%H:%M") == "09:45":
                row.append({
                    "type": "pause",
                    "label": "Grande pause"
                })

        timetable[day] = row

    return render(
        request,
        "timetable/timetable_grid.html",
        {
            "timetable": timetable,
            "weeks": weeks,
            "selected_week": selected_week,
            "sections": sections,
            "selected_section_id": str(section_id) if section_id else "",
        }
    )


@login_required
def course_create(request):
    if request.method == "POST":
        form = CourseScheduleForm(
            request.POST,
            user=request.user
        )

        if form.is_valid():
            course = form.save(commit=False)
            course.created_by = request.user
            course.save()

            messages.success(request, "Cours ajouté avec succès.")
            return redirect("timetable:timetable_grid")

    else:
        form = CourseScheduleForm(user=request.user)

    return render(
        request,
        "timetable/course_form.html",
        {"form": form}
    )


@login_required
def course_update(request, pk):
    course = get_object_or_404(
        CourseSchedule.objects.select_related(
            "week",
            "subject",
            "subject__level",
            "subject__level__section",
        ),
        pk=pk
    )

    if not user_can_manage_course(request.user, course):
        messages.error(request, "Accès refusé.")
        return redirect("timetable:timetable_grid")

    if course.week.is_locked:
        messages.error(request, "Cette semaine est verrouillée.")
        return redirect("timetable:timetable_grid")

    if request.method == "POST":
        form = CourseScheduleForm(
            request.POST,
            instance=course,
            user=request.user
        )

        if form.is_valid():
            form.save()
            messages.success(request, "Cours modifié avec succès.")
            return redirect("timetable:timetable_grid")

    else:
        form = CourseScheduleForm(
            instance=course,
            user=request.user
        )

    return render(
        request,
        "timetable/course_form.html",
        {"form": form}
    )


@login_required
def course_delete(request, pk):
    course = get_object_or_404(
        CourseSchedule.objects.select_related(
            "week",
            "subject",
            "subject__level",
            "subject__level__section",
        ),
        pk=pk
    )

    if not user_can_manage_course(request.user, course):
        messages.error(request, "Accès refusé.")
        return redirect("timetable:timetable_grid")

    if course.week.is_locked:
        messages.error(request, "Cette semaine est verrouillée.")
        return redirect("timetable:timetable_grid")

    if request.method == "POST":
        course.delete()
        messages.success(request, "Cours supprimé.")
        return redirect("timetable:timetable_grid")

    return render(
        request,
        "timetable/confirm_delete.html",
        {"object": course}
    )


@login_required
def week_list(request):
    weeks = AcademicWeek.objects.all().order_by("-start_date")

    return render(
        request,
        "timetable/week_list.html",
        {"weeks": weeks}
    )


@login_required
@user_passes_test(is_admin)
def week_create(request):
    if request.method == "POST":
        form = AcademicWeekForm(request.POST)

        if form.is_valid():
            form.save()
            messages.success(request, "Semaine ajoutée avec succès.")
            return redirect("timetable:week_list")

    else:
        form = AcademicWeekForm()

    return render(
        request,
        "timetable/week_form.html",
        {"form": form}
    )


@login_required
@user_passes_test(is_admin)
def week_update(request, pk):
    week = get_object_or_404(AcademicWeek, pk=pk)

    if request.method == "POST":
        form = AcademicWeekForm(
            request.POST,
            instance=week
        )

        if form.is_valid():
            form.save()
            messages.success(request, "Semaine modifiée.")
            return redirect("timetable:week_list")

    else:
        form = AcademicWeekForm(instance=week)

    return render(
        request,
        "timetable/week_form.html",
        {"form": form}
    )


@login_required
@user_passes_test(is_admin)
def week_delete(request, pk):
    week = get_object_or_404(AcademicWeek, pk=pk)

    if request.method == "POST":
        week.delete()
        messages.success(request, "Semaine supprimée.")
        return redirect("timetable:week_list")

    return render(
        request,
        "timetable/confirm_delete.html",
        {"object": week}
    )


@login_required
def change_password(request):
    if request.method == "POST":
        form = PasswordChangeForm(
            request.user,
            request.POST
        )

        if form.is_valid():
            user = form.save()

            update_session_auth_hash(
                request,
                user
            )

            messages.success(request, "Mot de passe modifié.")
            return redirect("timetable:dashboard")

    else:
        form = PasswordChangeForm(request.user)

    return render(
        request,
        "timetable/change_password.html",
        {"form": form}
    )


@login_required
@user_passes_test(is_admin)
def reset_chief_password(request, user_id):
    user = get_object_or_404(User, pk=user_id)

    user.set_password("1234")
    user.save()

    messages.success(
        request,
        f"Mot de passe de {user.username} réinitialisé à 1234."
    )

    return redirect("timetable:chief_list")


def format_course(course):
    if not course:
        return "Libre"

    return (
        f"{course.subject.name}\n"
        f"{course.teacher}\n"
        f"{course.subject.level.name}\n"
        f"Salle : {course.room.name}"
    )


@login_required
def export_timetable_excel(request):
    days = [
        "LUNDI",
        "MARDI",
        "MERCREDI",
        "JEUDI",
        "VENDREDI",
        "SAMEDI",
    ]

    columns = [
        "Jour",
        "07h30 - 09h30",
        "09h30 - 09h45",
        "09h45 - 12h45",
        "12h45 - 14h00",
        "14h00 - 17h00",
    ]

    selected_week = get_selected_week(request)

    if request.user.is_staff or request.user.is_superuser:
        sections = Section.objects.select_related(
            "department",
            "chief"
        ).order_by(
            "department__name",
            "name"
        )
    else:
        sections = Section.objects.filter(
            id=request.user.managed_section.id
        ).select_related(
            "department",
            "chief"
        )

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    for section in sections:
        sheet_name = section.name[:31]
        sheet = workbook.create_sheet(title=sheet_name)

        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
        sheet.page_margins.left = 0.3
        sheet.page_margins.right = 0.3
        sheet.page_margins.top = 0.5
        sheet.page_margins.bottom = 0.5

        logo_path = os.path.join(
            settings.BASE_DIR,
            "timetable",
            "static",
            "timetable",
            "images",
            "logo.png"
        )

        if os.path.exists(logo_path):
            logo = Image(logo_path)
            logo.width = 80
            logo.height = 80
            sheet.add_image(logo, "A1")

        sheet.merge_cells("B1:F1")
        sheet["B1"] = "EMPLOI DU TEMPS HEBDOMADAIRE"
        sheet["B1"].font = Font(bold=True, size=16)
        sheet["B1"].alignment = Alignment(horizontal="center")

        sheet.merge_cells("B2:F2")
        sheet["B2"] = f"Département : {section.department.name}"
        sheet["B2"].font = Font(bold=True)
        sheet["B2"].alignment = Alignment(horizontal="center")

        sheet.merge_cells("B3:F3")
        sheet["B3"] = f"Section : {section.name}"
        sheet["B3"].font = Font(bold=True)
        sheet["B3"].alignment = Alignment(horizontal="center")

        sheet.merge_cells("B4:F4")

        if selected_week:
            sheet["B4"] = f"Semaine : {selected_week.title} | {selected_week.start_date} au {selected_week.end_date}"
        else:
            sheet["B4"] = "Semaine : non définie"

        sheet["B4"].alignment = Alignment(horizontal="center")

        start_row = 6

        for col_num, header in enumerate(columns, 1):
            cell = sheet.cell(row=start_row, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="9C6B30")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        schedules = CourseSchedule.objects.filter(
            subject__level__section=section
        ).select_related(
            "week",
            "subject",
            "subject__level",
            "teacher",
            "room",
            "time_slot"
        )

        if selected_week:
            schedules = schedules.filter(week=selected_week)

        schedules = schedules.order_by(
            "day",
            "time_slot__start_time",
            "subject__level__name"
        )

        row = start_row + 1

        for day in days:
            sheet.cell(row=row, column=1).value = day

            day_courses = schedules.filter(day=day)

            slot_1 = day_courses.filter(
                time_slot__start_time="07:30"
            ).first()

            slot_2 = day_courses.filter(
                time_slot__start_time="09:45"
            ).first()

            slot_3 = day_courses.filter(
                time_slot__start_time="14:00"
            ).first()

            sheet.cell(row=row, column=2).value = format_course(slot_1)
            sheet.cell(row=row, column=3).value = "PETITE PAUSE"
            sheet.cell(row=row, column=4).value = format_course(slot_2)
            sheet.cell(row=row, column=5).value = "GRANDE PAUSE"
            sheet.cell(row=row, column=6).value = format_course(slot_3)

            row += 1

        signature_row = row + 3

        sheet.merge_cells(
            start_row=signature_row,
            start_column=5,
            end_row=signature_row,
            end_column=6
        )

        sheet.cell(row=signature_row, column=5).value = "Le Chef de section"
        sheet.cell(row=signature_row, column=5).alignment = Alignment(horizontal="center")

        sheet.merge_cells(
            start_row=signature_row + 3,
            start_column=5,
            end_row=signature_row + 3,
            end_column=6
        )

        if section.chief:
            chief_name = f"{section.chief.first_name} {section.chief.last_name}"
        else:
            chief_name = "Non défini"

        sheet.cell(row=signature_row + 3, column=5).value = chief_name
        sheet.cell(row=signature_row + 3, column=5).font = Font(bold=True)
        sheet.cell(row=signature_row + 3, column=5).alignment = Alignment(horizontal="center")

        thin = Side(border_style="thin", color="000000")

        for rows in sheet.iter_rows(
            min_row=start_row,
            max_row=row - 1,
            min_col=1,
            max_col=6
        ):
            for cell in rows:
                cell.border = Border(
                    top=thin,
                    left=thin,
                    right=thin,
                    bottom=thin
                )
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )

        for col in range(1, 7):
            sheet.column_dimensions[get_column_letter(col)].width = 25

        for r in range(start_row + 1, row):
            sheet.row_dimensions[r].height = 75

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = (
        'attachment; filename="emplois_du_temps_par_section.xlsx"'
    )

    workbook.save(response)

    return response


@login_required
def teacher_list(request):
    teachers = Teacher.objects.all().order_by(
        "last_name",
        "first_name"
    )

    return render(
        request,
        "timetable/teacher_list.html",
        {"teachers": teachers}
    )


@login_required
@user_passes_test(is_admin)
def teacher_create(request):
    if request.method == "POST":
        form = TeacherForm(request.POST)

        if form.is_valid():
            form.save()
            messages.success(request, "Enseignant ajouté avec succès.")
            return redirect("timetable:teacher_list")

    else:
        form = TeacherForm()

    return render(
        request,
        "timetable/teacher_form.html",
        {"form": form}
    )


@login_required
@user_passes_test(is_admin)
def teacher_update(request, pk):
    teacher = get_object_or_404(Teacher, pk=pk)

    if request.method == "POST":
        form = TeacherForm(
            request.POST,
            instance=teacher
        )

        if form.is_valid():
            form.save()
            messages.success(request, "Enseignant modifié.")
            return redirect("timetable:teacher_list")

    else:
        form = TeacherForm(instance=teacher)

    return render(
        request,
        "timetable/teacher_form.html",
        {"form": form}
    )


@login_required
@user_passes_test(is_admin)
def teacher_delete(request, pk):
    teacher = get_object_or_404(Teacher, pk=pk)

    if request.method == "POST":
        teacher.delete()
        messages.success(request, "Enseignant supprimé.")
        return redirect("timetable:teacher_list")

    return render(
        request,
        "timetable/confirm_delete.html",
        {"object": teacher}
    )


@login_required
def department_list(request):
    departments = Department.objects.all().order_by("name")

    return render(
        request,
        "timetable/department_list.html",
        {"departments": departments}
    )


@login_required
@user_passes_test(is_admin)
def department_create(request):
    if request.method == "POST":
        form = DepartmentForm(request.POST)

        if form.is_valid():
            form.save()
            messages.success(request, "Département ajouté avec succès.")
            return redirect("timetable:department_list")

    else:
        form = DepartmentForm()

    return render(
        request,
        "timetable/department_form.html",
        {"form": form}
    )


@login_required
@user_passes_test(is_admin)
def department_update(request, pk):
    department = get_object_or_404(Department, pk=pk)

    if request.method == "POST":
        form = DepartmentForm(
            request.POST,
            instance=department
        )

        if form.is_valid():
            form.save()
            messages.success(request, "Département modifié.")
            return redirect("timetable:department_list")

    else:
        form = DepartmentForm(instance=department)

    return render(
        request,
        "timetable/department_form.html",
        {"form": form}
    )


@login_required
@user_passes_test(is_admin)
def department_delete(request, pk):
    department = get_object_or_404(Department, pk=pk)

    if request.method == "POST":
        department.delete()
        messages.success(request, "Département supprimé.")
        return redirect("timetable:department_list")

    return render(
        request,
        "timetable/confirm_delete.html",
        {"object": department}
    )


@login_required
def section_list(request):
    sections = Section.objects.select_related(
        "department",
        "chief"
    ).order_by(
        "department__name",
        "name"
    )

    return render(
        request,
        "timetable/section_list.html",
        {"sections": sections}
    )


@login_required
@user_passes_test(is_admin)
def section_create(request):
    if request.method == "POST":
        form = SectionForm(request.POST)

        if form.is_valid():
            form.save()
            messages.success(request, "Section ajoutée avec succès.")
            return redirect("timetable:section_list")

    else:
        form = SectionForm()

    return render(
        request,
        "timetable/section_form.html",
        {"form": form}
    )


@login_required
@user_passes_test(is_admin)
def section_update(request, pk):
    section = get_object_or_404(Section, pk=pk)

    if request.method == "POST":
        form = SectionForm(
            request.POST,
            instance=section
        )

        if form.is_valid():
            form.save()
            messages.success(request, "Section modifiée.")
            return redirect("timetable:section_list")

    else:
        form = SectionForm(instance=section)

    return render(
        request,
        "timetable/section_form.html",
        {"form": form}
    )


@login_required
@user_passes_test(is_admin)
def section_delete(request, pk):
    section = get_object_or_404(Section, pk=pk)

    if request.method == "POST":
        section.delete()
        messages.success(request, "Section supprimée.")
        return redirect("timetable:section_list")

    return render(
        request,
        "timetable/confirm_delete.html",
        {"object": section}
    )


@login_required
def level_list(request):
    if request.user.is_staff or request.user.is_superuser:
        levels = Level.objects.select_related(
            "section",
            "section__department"
        )
    else:
        levels = Level.objects.filter(
            section=request.user.managed_section
        ).select_related(
            "section",
            "section__department"
        )

    levels = levels.order_by(
        "section__department__name",
        "section__name",
        "name"
    )

    return render(
        request,
        "timetable/level_list.html",
        {"levels": levels}
    )


@login_required
@user_passes_test(is_admin)
def level_create(request):
    if request.method == "POST":
        form = LevelForm(request.POST)

        if form.is_valid():
            form.save()
            messages.success(request, "Niveau ajouté avec succès.")
            return redirect("timetable:level_list")

    else:
        form = LevelForm()

    return render(
        request,
        "timetable/level_form.html",
        {"form": form}
    )


@login_required
@user_passes_test(is_admin)
def level_update(request, pk):
    level = get_object_or_404(Level, pk=pk)

    if request.method == "POST":
        form = LevelForm(
            request.POST,
            instance=level
        )

        if form.is_valid():
            form.save()
            messages.success(request, "Niveau modifié.")
            return redirect("timetable:level_list")

    else:
        form = LevelForm(instance=level)

    return render(
        request,
        "timetable/level_form.html",
        {"form": form}
    )


@login_required
@user_passes_test(is_admin)
def level_delete(request, pk):
    level = get_object_or_404(Level, pk=pk)

    if request.method == "POST":
        level.delete()
        messages.success(request, "Niveau supprimé.")
        return redirect("timetable:level_list")

    return render(
        request,
        "timetable/confirm_delete.html",
        {"object": level}
    )


@login_required
def room_list(request):
    rooms = Room.objects.all().order_by("name")

    return render(
        request,
        "timetable/room_list.html",
        {"rooms": rooms}
    )


@login_required
@user_passes_test(is_admin)
def room_create(request):
    if request.method == "POST":
        form = RoomForm(request.POST)

        if form.is_valid():
            form.save()
            messages.success(request, "Salle ajoutée avec succès.")
            return redirect("timetable:room_list")

    else:
        form = RoomForm()

    return render(
        request,
        "timetable/room_form.html",
        {"form": form}
    )


@login_required
@user_passes_test(is_admin)
def room_update(request, pk):
    room = get_object_or_404(Room, pk=pk)

    if request.method == "POST":
        form = RoomForm(
            request.POST,
            instance=room
        )

        if form.is_valid():
            form.save()
            messages.success(request, "Salle modifiée.")
            return redirect("timetable:room_list")

    else:
        form = RoomForm(instance=room)

    return render(
        request,
        "timetable/room_form.html",
        {"form": form}
    )


@login_required
@user_passes_test(is_admin)
def room_delete(request, pk):
    room = get_object_or_404(Room, pk=pk)

    if request.method == "POST":
        room.delete()
        messages.success(request, "Salle supprimée.")
        return redirect("timetable:room_list")

    return render(
        request,
        "timetable/confirm_delete.html",
        {"object": room}
    )


@login_required
def subject_list(request):
    if request.user.is_staff or request.user.is_superuser:
        subjects = Subject.objects.select_related(
            "level",
            "level__section",
            "level__section__department"
        )
    else:
        subjects = Subject.objects.filter(
            level__section=request.user.managed_section
        ).select_related(
            "level",
            "level__section",
            "level__section__department"
        )

    subjects = subjects.order_by(
        "level__section__department__name",
        "level__section__name",
        "level__name",
        "name"
    )

    return render(
        request,
        "timetable/subject_list.html",
        {"subjects": subjects}
    )


@login_required
def subject_create(request):
    if request.method == "POST":
        form = SubjectForm(
            request.POST,
            user=request.user
        )

        if form.is_valid():
            subject = form.save(commit=False)
            subject.created_by = request.user
            subject.save()

            messages.success(request, "Matière ajoutée.")
            return redirect("timetable:subject_list")

    else:
        form = SubjectForm(user=request.user)

    return render(
        request,
        "timetable/subject_form.html",
        {"form": form}
    )


@login_required
def subject_update(request, pk):
    subject = get_object_or_404(
        Subject.objects.select_related(
            "level",
            "level__section"
        ),
        pk=pk
    )

    if not user_can_manage_subject(request.user, subject):
        messages.error(request, "Accès refusé.")
        return redirect("timetable:subject_list")

    if request.method == "POST":
        form = SubjectForm(
            request.POST,
            instance=subject,
            user=request.user
        )

        if form.is_valid():
            form.save()
            messages.success(request, "Matière modifiée.")
            return redirect("timetable:subject_list")

    else:
        form = SubjectForm(
            instance=subject,
            user=request.user
        )

    return render(
        request,
        "timetable/subject_form.html",
        {"form": form}
    )


@login_required
def subject_delete(request, pk):
    subject = get_object_or_404(
        Subject.objects.select_related(
            "level",
            "level__section"
        ),
        pk=pk
    )

    if not user_can_manage_subject(request.user, subject):
        messages.error(request, "Accès refusé.")
        return redirect("timetable:subject_list")

    if request.method == "POST":
        subject.delete()
        messages.success(request, "Matière supprimée.")
        return redirect("timetable:subject_list")

    return render(
        request,
        "timetable/confirm_delete.html",
        {"object": subject}
    )


@login_required
@user_passes_test(is_admin)
def chief_list(request):
    chiefs = User.objects.filter(
        managed_section__isnull=False
    ).select_related(
        "managed_section",
        "managed_section__department"
    )

    return render(
        request,
        "timetable/chief_list.html",
        {"chiefs": chiefs}
    )


@login_required
@user_passes_test(is_admin)
def chief_create(request):
    if request.method == "POST":
        form = ChiefUserForm(request.POST)

        if form.is_valid():
            form.save()
            messages.success(request, "Chef créé avec succès.")
            return redirect("timetable:chief_list")

    else:
        form = ChiefUserForm(
            initial={"password": "1234"}
        )

    return render(
        request,
        "timetable/chief_form.html",
        {"form": form}
    )